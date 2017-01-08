from openpyxl import load_workbook
from openpyxl import Workbook
from bs4 import BeautifulSoup
from datetime import date
from operator import itemgetter
import requests
import time
import sys
import datetime
import operator
import numpy as np
import csv


def getIDs():
	wb = load_workbook(filename = 'ESPNplayerIDs.xlsx') #load the workbook  
	ws = wb['Names and IDs'] #grab the worksheet from the workbook
	playerIDs = {} #create a dictionary {playeer name -> player ID}

	#loop through all 464 players that are in the workbook put them with their ID's in the dictionary
	for row in range(1,464):
	    playerCell = "A" + str(row)
	    idCell = "B" + str(row)
	    playerArray = ws[playerCell].value.lower().split("-")
	    player = playerArray[0] + " " + playerArray[1]
	    iD = int(ws[idCell].value)
	    playerIDs[player] = iD
	return playerIDs


def getStats(playerID, HomeAwayOrOpp):
	base_url = 'http://espn.go.com/nba/player/gamelog/_/id/'
	page = base_url + str(playerID)
	print(page)
	r = requests.get(page)
	#soup = BeautifulSoup(r.text, 'lxml')
	soup = BeautifulSoup(r.text)
	table = soup.find("table",{"class" : "tablehead"})
	data = [[stat.get_text() for stat in t.find_all("td")] for t in table.find_all("tr")]
	for array in data: 
		if len(array) != 17:
			data.remove(array)
	pnts = 0
	rebs = 0
	asst = 0
	block = 0
	stl = 0
	to = 0
	count = 0
	dbldbl = 0
	tpldbl = 0
	three = 0

	for array in data:
		if len(array) < 2:
			continue
		if HomeAwayOrOpp == 'home':
			if array[1][:2] == 'vs':
				curPnts, curRebs, curAsst, curBlock, curStl, curTo, curThree, curTpldbl, curDbldbl, curCount = getFantasyStats(array)
				count += curCount
				three = curThree
				rebs += curRebs
				asst += curAsst
				block += curBlock
				stl += curStl
				to += curTo
				pnts += curPnts
					
		elif HomeAwayOrOpp == 'away':
			if array[1][0]== '@':
				curPnts, curRebs, curAsst, curBlock, curStl, curTo, curThree, curTpldbl, curDbldbl, curCount = getFantasyStats(array)
				count += curCount
				three = curThree
				rebs += curRebs
				asst += curAsst
				block += curBlock
				stl += curStl
				to += curTo
				pnts += curPnts
		else:
			if (array[1][2:]) == HomeAwayOrOpp:
				curPnts, curRebs, curAsst, curBlock, curStl, curTo, curThree, curTpldbl, curDbldbl, curCount = getFantasyStats(array)
				count += curCount
				three = curThree
				rebs += curRebs
				asst += curAsst
				block += curBlock
				stl += curStl
				to += curTo
				pnts += curPnts
	return getFantasyPoints(pnts, rebs, asst, block, stl, to, three, tpldbl, dbldbl, count)			 
				  
def getFantasyStats(array):
	count = 1.0
	three = float(str(array[-11]).split("-")[0])
	rebs = float(array[-7])
	asst = float(array[-6])
	block = float(array[-5])
	stl = float(array[-4])
	to = float(array[-3])
	pnts = float(array[-1])
	stat = [float(array[-1]),float(array[-4]),float(array[-5]),float(array[-6]),float(array[-7])]
	countdouble = 0
	#test for triple doubles and double doubles
	tpldbl = 0
	dbldbl = 0
	for s in stat:
		if s >= 10:
			countdouble += 1.0
	if countdouble == 2:
		dbldbl = 1.0
	elif countdouble > 2:
		tpldbl = 1.0
	return (pnts, rebs, asst, block, stl, to, three, tpldbl, dbldbl, count)

def getFantasyPoints(pnts, rebs, asst, block, stl, to, three, tpldbl, dbldbl, count):
	if count == 0:
		return 0

	pnts = pnts/count
	rebs = rebs/count
	asst = asst/count
	block = block/count
	stl = stl/count
	to = to/count
	three = three/count
	dbldbl = dbldbl/count
	tpldbl = tpldbl/count
	fantasypnts = 0.5 * three
	fantasypnts += 1.5*dbldbl + 3*tpldbl
	fantasypnts += pnts + 1.25*rebs + 1.5*asst + 2*stl + 2*block
	fantasypnts -= 0.5*to
	#print("Average total fantasy points: " + str(fantasypnts))
	return fantasypnts




